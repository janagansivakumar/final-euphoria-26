from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

excel_file = "participants.xlsx"

# Create workbook if not exists
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(["User Code", "Name", "College", "Department", "Year", "Email", "Phone", "Events", "Team Members"])
    wb.save(excel_file)

def generate_user_code():
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Participants"]
    # Subtract 1 for header row
    count = ws.max_row - 1
    next_number = count + 1
    return f"EUP-{next_number:04d}"   # e.g. EUP-0001, EUP-0002

@app.route("/")
def form():
    return render_template("register.html")

@app.route("/register", methods=["POST"])
def register():
    fullName = request.form.get("fullName")
    collegeName = request.form.get("collegeName")
    department = request.form.get("department")
    yearOfStudy = request.form.get("yearOfStudy")
    email = request.form.get("email")
    phone = request.form.get("phone")
    events = ", ".join(request.form.getlist("events"))

    # Collect team members if group event selected
    team_members = []
    for i in range(1, 11):  # allow up to 10 members
        member = request.form.get(f"member{i}")
        if member:
            team_members.append(member)
    team_members_str = ", ".join(team_members) if team_members else "N/A"

    userCode = generate_user_code()

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Participants"]
    ws.append([userCode, fullName, collegeName, department, yearOfStudy, email, phone, events, team_members_str])
    wb.save(excel_file)

    return redirect(url_for("success", userCode=userCode, fullName=fullName, events=events, team=team_members_str))

@app.route("/success")
def success():
    userCode = request.args.get("userCode")
    fullName = request.args.get("fullName")
    events = request.args.get("events")
    team = request.args.get("team")
    return render_template("success.html", userCode=userCode, fullName=fullName, events=events, team=team)

if __name__ == "__main__":
    app.run(debug=True)
